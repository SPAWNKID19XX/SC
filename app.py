#Python 3.12.3
import os, csv
from datetime import datetime
from flask import Flask, render_template,redirect,url_for, request 
import openpyxl.workbook
from form import FormSubscribe
from secret import SECRET_KEY
from flask_mail import Mail, Message
import my_secret_data
from twilio.rest import Client
from openpyxl import Workbook, load_workbook
import gspread
from google.oauth2.service_account import Credentials
import countries
from pythonmonkey import require as js_require
import json

countries_list = countries.countries_list

app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = my_secret_data.MAIL_USERNAME
app.config['MAIL_PASSWORD'] = my_secret_data.MAIL_PASSWORD
app.config['MAIL_DEFAULT_SENDER'] = my_secret_data.MAIL_SENDER


mail=Mail(app)


account_sid = my_secret_data.ACCOUNT_SID
auth_token = my_secret_data.AUTH_TOKEN
client = Client(account_sid, auth_token)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
users_csv_file_path = os.path.join(BASE_DIR, 'users.csv')
credentials_file_path = os.path.join(BASE_DIR, 'credentials.json')
users_xlsx_file_path = os.path.join(BASE_DIR, 'users.xlsx')
field_names = ['Name', 'Country', 'Whatsapp']


@app.route('/', methods=['GET', 'POST'])
def index():
    csv_to_exel()
    form = FormSubscribe()
    if form.validate_on_submit():
        full_name = form.full_name.data
        country = form.countries.data
        wtsapp = form.wtsapp.data
        country_code = get_couontry_code(country)
        country_code_numer = f"({country_code}) {wtsapp}"
        now = datetime.now()
        timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
        success_msg = f"Thank you for subscribing, {full_name} from {country}! We will contact you at {country_code_numer}."
        print(success_msg)
        if request.method == 'POST':
            send_email(full_name, country, wtsapp)
            
            new_rec_xl = [timestamp , full_name, country, f'{country_code}{wtsapp}']
            #res = add_new_rec_to_xlsx(new_rec_xl)
            if add_new_rec_to_xlsx(new_rec_xl):
                return render_template('user_already_exist_template.html', country_code=country_code, wtsapp=wtsapp )
            return redirect(url_for('accept'))
    return render_template('index.html', form=form)


@app.route('/accept')
def accept():
    return render_template('accept_appointment_template.html')

@app.route('/gold_print_mentoria')
def gold_print_mentoria():
    return render_template('gold_print_mentoria.html')

@app.route('/gold_print_form', methods=['GET', 'POST'])
def gold_print_form():
    form = FormSubscribe()
    if form.validate_on_submit():
        full_name = form.full_name.data
        country = form.countries.data
        wtsapp = form.wtsapp.data
        country_code = get_couontry_code(country)
        country_code_numer = f"({country_code}) {wtsapp}"
        now = datetime.now()
        timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
        success_msg = f"Thank you for subscribing, {full_name} from {country}! We will contact you at {country_code_numer}."
        print(success_msg)
        if request.method == 'POST':
            send_email(full_name, country, wtsapp)
            
            new_rec_xl = [timestamp , full_name, country, f'{country_code}{wtsapp}']
            #res = add_new_rec_to_xlsx(new_rec_xl)
            if add_new_rec_to_xlsx(new_rec_xl):
                return render_template('user_already_exist_template.html', country_code=country_code, wtsapp=wtsapp )
            return redirect(url_for('accept'))
    return render_template('form_page.html', form=form)


def get_json_new_rec(new_rec:list):
    result_json = {
        'data_time' : new_rec[0],
        'name' : new_rec[1],
        'countrie' : new_rec[2],
        'whatsapp_number' :new_rec[3],
    }
    #return json.dumps(result_json)
    
    scopes = [
    "https://www.googleapis.com/auth/spreadsheets"
]
    creds = Credentials.from_service_account_file('credentials.json', scopes=scopes)
    client = gspread.authorize(creds)
    SHEET_ID = "15buaxLcT9t8guvCbNpM_Ia-BHCuU50eTiPejPvvU2PM"
    sheet = client.open_by_key(SHEET_ID)
    values_list = sheet.sheet1.append_row(new_rec)
    

def create_exel_file_users():
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'users'
    sheet.append(field_names)
    wb.save(users_xlsx_file_path)


def is_exist_whatsapp_number(whatsapp):
    '''
    Check if var whatsapp exist in exel.sheet
    '''

    wb = load_workbook(users_xlsx_file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value == whatsapp:
                return True
    return False


def add_new_rec_to_xlsx(new_rec_xl):
    '''
    create a exel file if it does not exist, 
    and add a new rec in last possition
    '''
    if not os.path.exists(users_xlsx_file_path):        
        create_exel_file_users()
    else:
        wb = load_workbook(users_xlsx_file_path)
        if 'users' in wb.sheetnames:
            sheet = wb['users']
        else:
            sheet = wb.create_sheet('users')
    
    if not is_exist_whatsapp_number(new_rec_xl[-1]):
        sheet.append(new_rec_xl)
        wb.save(users_xlsx_file_path)
        get_json_new_rec(new_rec_xl)
    else:
        print(f"{new_rec_xl[-1]} ALREADY EXISTS!!!")
        return True

    
def csv_to_exel():
    '''
    if csv exist it pass all rec's into exel file 
    and if number already exist, this rec is ignored.
    to finish with csv it remove a file
    '''
    if not os.path.exists(users_xlsx_file_path):        
        create_exel_file_users()

    if os.path.exists(users_csv_file_path):
        print('CSV Exist')
        wb = load_workbook(users_xlsx_file_path)
        sheet = wb['users']

        with open(users_csv_file_path, 'r') as csv_file:
            reader = csv.reader(csv_file)
            now = datetime.now()
            timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
            for row in reader:
                row.insert(0,timestamp)
                digits = row[-1] 
                if digits[1:].isdigit() and not is_exist_whatsapp_number(row[-1]):
                    sheet.append(row)
                else:
                    print(row[-1],'already exist in list')
            os.remove(users_csv_file_path)
        wb.save(users_xlsx_file_path)


def send_email(name, country, whatsapp):
    recipient = my_secret_data.MAIL_SENDER
    subject = "New User Subscribe"
    message_body = f"""
    Lead para amentoria GOLDPRINT
    Dados de contacto:
    Name :{name}
    Country :{country} 
    Watsapp {whatsapp}
    """

    message = Message(subject=subject,
                        recipients=[recipient],
                        body=message_body)
    
    try:
        mail.send(message)
        print('Email sent successfully!') 
    except Exception as e:
        print('DOES NOT')
        print(str(e))


def send_msg_whatsapp(name, whatsapp):
    message = client.messages.create(
    from_='whatsapp:+14155238886',
    body=f'Hello {name}.\n we glad then you wont visit us',
    to=[f'whatsapp:{whatsapp}']
    )
    print(message.sid)


def get_couontry_code(country):
    return countries_list[country]

if __name__ == '__main__':
    app.run(debug=True)
